import json
import logging
import re
import sys
from datetime import datetime, timedelta
from io import BytesIO
from pprint import pprint

import boto3
import django
import twd97

django.setup()
from bucket.models import Bucket
from bucketRecord.models import BucketRecord
from openpyxl import load_workbook


log = logging.getLogger('')
log.setLevel(logging.DEBUG)
format = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(format)
log.addHandler(ch)

parselogger = logging.getLogger('parse')
sheetlogger = logging.getLogger('sheet')
progresslogger = logging.getLogger('progress')

NO_DATA = -1

Bucket.objects.all().delete()

s3 = boto3.resource('s3')
s3_client = boto3.client('s3')
file_list = list()
village_list = [];
weeknum = int(datetime.now().strftime('%U'))
this_year_str = datetime.now().strftime('%Y')
accept_weeks = []

accept_weeks.append(str(weeknum))
accept_weeks.append(str(weeknum - 1))


for key in s3.Bucket('dengue-report-source').objects.all():
    if key.key.endswith(".xlsx"):
        if len(key.key.split("/")) < 2:
            continue
        city = key.key.split("/")[0]
        file_name = key.key.split("/")[1]
        file_list.append({
            "city": city,
            "file_name": file_name,
            "file_key": key.key
        })

bucket_dict = dict()
survey_dict = dict()
for file_dict in file_list:

    s3_obj = s3.Object('dengue-report-source', file_dict['file_key'])
    wb = load_workbook(filename=BytesIO(s3_obj.get()['Body'].read()), read_only=True)
    parselogger.info(file_dict['file_name'])
    city = file_dict['city']
    for sheet_name in wb.get_sheet_names():
        sheet_name_match = re.search(r'(\d+)(年)?第(\d+)(週|周)', sheet_name)
        if sheet_name == '誘卵桶資訊':
            sheetlogger.info(sheet_name)
            ws = wb['誘卵桶資訊']
            progresslogger.info('Bucket handle begin')
            for row in range(3, ws.max_row+1):
                bucket_id = ws['A' + str(row)].value
                if bucket_id == None:
                    continue

                bucket_x = ws['B' + str(row)].value
                bucket_y = ws['C' + str(row)].value
                if bucket_x == None or bucket_y == None:
                    continue
                try:
                    bucket_x = float(bucket_x)
                    bucket_y = float(bucket_y)
                except:
                    continue

                bucket_address = ws['D' + str(row)].value if isinstance(ws['D' + str(row)].value, str) == True else '無'
                bucket_note = ws['E' + str(row)].value if isinstance(ws['E' + str(row)].value, str) == True else '無'
                bucket_lat, bucket_lng = twd97.towgs84(bucket_x, bucket_y)
                bucket_dict[bucket_id] = {
                    'bucket_lat': bucket_lat if isinstance(bucket_lat, float) == True else NO_DATA,
                    'bucket_lng': bucket_lng if isinstance(bucket_lng, float) == True else NO_DATA,
                    # 'bucket_address': bucket_address,
                    # 'bucket_note': bucket_note
                }
                Bucket(id=bucket_id,
                       ws84_x=bucket_x,
                       ws84_y=bucket_y,
                       address=bucket_address,
                       note=bucket_note,
                       lng=bucket_lng,
                       lat=bucket_lat,
                       point='POINT(%f %f)' % (bucket_x, bucket_y),
                       ).save()
            progresslogger.info('finish ' + str(ws.max_row + 1) + ' buckets')

        elif sheet_name_match and sheet_name_match.group(3) in accept_weeks and sheet_name_match.group(1) == this_year_str:
            ws = wb[sheet_name]
            sheetlogger.info(sheet_name)
            progresslogger.info('record handle begin')
            for row in range(3, ws.max_row+1):
                survey_date = ws['A' + str(row)].value
                bucket_id = ws['B' + str(row)].value
                if isinstance(survey_date, datetime) == False or \
                        bucket_dict.get(bucket_id) == None:
                    break

                area = ws['C' + str(row)].value
                village = ws['D' + str(row)].value
                if '里' not in village:
                    village = village + '里'

                egg_num = ws['E' + str(row)].value if isinstance(ws['E' + str(row)].value, int) == True else NO_DATA
                if egg_num == 0:
                    egypt_egg_num = 0
                    white_egg_num = 0
                else:
                    egypt_egg_num = ws['F' + str(row)].value if isinstance(ws['F' + str(row)].value, int) == True else NO_DATA
                    white_egg_num = ws['G'+ str(row)].value if isinstance(ws['G' + str(row)].value, int) == True else NO_DATA

                larvae_num = ws['H' + str(row)].value if isinstance(ws['H' + str(row)].value, int) == True else NO_DATA
                if larvae_num == 0:
                    egypt_larvae_num = 0
                    white_larvae_num = 0
                else:
                    egypt_larvae_num = ws['I' + str(row)].value if isinstance(ws['I' + str(row)].value, int) == True else NO_DATA
                    white_larvae_num = ws['J' + str(row)].value if isinstance(ws['J' + str(row)].value, int) == True else NO_DATA
                survey_note = ws['K' + str(row)].value if isinstance(ws['K' + str(row)].value, str) == True else '無'

                if([city, area, village] not in village_list):
                    village_list.append([city, area, village])
                
                dup_records_num = len(BucketRecord.objects.filter(
                    bucket_id=bucket_id,
                    investigate_date=survey_date,
                    county=city
                ))

                if dup_records_num > 1:
                    warning_str = 'duplicated records:' + str(bucket_id) + str(survey_date) + city
                    progresslogger.warning(warning_str)
                    record_exist = True
                elif dup_records_num == 1:
                    record_exist = True
                elif dup_records_num == 0:
                    record_exist = False
                
                if not record_exist:
                    progresslogger.info(
                        'save object:' + str(bucket_id) + ',' + str(survey_date))
                    BucketRecord(
                        bucket_id=bucket_id,
                        investigate_date=survey_date,
                        county=city,
                        town=area,
                        village=village,
                        # 卵數、埃及孵化卵數、白線孵化卵數
                        egg_count=egg_num,
                        egypt_egg_count=egypt_egg_num,
                        white_egg_count=white_egg_num,
                        # 孑孓、埃及幼蟲、白線幼蟲
                        larvae_count=larvae_num,
                        egypt_larvae_count=egypt_larvae_num,
                        white_larvae_count=white_larvae_num,
                        note=survey_note,
                    ).save()
                    
            progresslogger.info('finish ' + str(ws.max_row + 1) + ' records')
    wb.close()
