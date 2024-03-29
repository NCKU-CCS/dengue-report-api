import boto3
import json
import re
import twd97
import django

from io import BytesIO
from datetime import datetime, timedelta
from pprint import pprint
from openpyxl import load_workbook

django.setup()

from bucket.models import Bucket
from bucketRecord.models import BucketRecord
from bucketStatistics.models import BucketStatistics


# def generate_week_str(now_week_start, week_ago_num):
#     week_ago_start = now_week_start - timedelta(days=7 * week_ago_num)
#     week_ago_end = now_week_start - timedelta(days=1 * week_ago_num)
#     return '%s~%s' % (
#         week_ago_start.strftime("%Y-%m-%d"), week_ago_end.strftime("%Y-%m-%d"))

NO_DATA = -1

Bucket.objects.all().delete()
BucketRecord.objects.all().delete();

s3 = boto3.resource('s3')
s3_client = boto3.client('s3')
file_list = list()
village_list = [];

for key in s3.Bucket('dengue-report-source').objects.all():
    if key.key.endswith(".xlsx"):
        if len(key.key.split("/")) < 2:
            continue
        city = key.key.split("/")[0]
        file_name = key.key.split("/")[1]
        file_list.append({
            "city": city,
            "file_name": file_name,
            "file_key": key.key
        })

bucket_dict = dict()
survey_dict = dict()
for file_dict in file_list:
    # s3_client.download_file(
        # 'dengue-report-source',
        # file_dict['file_key'],
        # file_dict['file_name']
    # )

    # wb = load_workbook(file_dict['file_name'], read_only=True)
    s3_obj = s3.Object('dengue-report-source', file_dict['file_key'])
    wb = load_workbook(filename=BytesIO(s3_obj.get()['Body'].read()), read_only=True)
    print(file_dict['file_name'])
    city = file_dict['city']
    for sheet_name in wb.get_sheet_names():
        sheet_name_match = re.search(r'\d+(年)?第\d+(週|周)', sheet_name)
        if sheet_name == '誘卵桶資訊':
            ws = wb['誘卵桶資訊']
            for row in range(3, ws.max_row+1):
                bucket_id = ws['A' + str(row)].value
                if bucket_id == None:
                    continue

                bucket_x = ws['B' + str(row)].value
                bucket_y = ws['C' + str(row)].value
                if bucket_x == None or bucket_y == None:
                    continue
                try:
                    bucket_x = float(bucket_x)
                    bucket_y = float(bucket_y)
                except:
                    continue

                bucket_address = ws['D' + str(row)].value if isinstance(ws['D' + str(row)].value, str) == True else '無'
                bucket_note = ws['E' + str(row)].value if isinstance(ws['E' + str(row)].value, str) == True else '無'
                bucket_lat, bucket_lng = twd97.towgs84(bucket_x, bucket_y)
                bucket_dict[bucket_id] = {
                    'bucket_lat': bucket_lat if isinstance(bucket_lat, float) == True else NO_DATA,
                    'bucket_lng': bucket_lng if isinstance(bucket_lng, float) == True else NO_DATA,
                    # 'bucket_address': bucket_address,
                    # 'bucket_note': bucket_note
                }
                Bucket(id=bucket_id,
                       ws84_x=bucket_x,
                       ws84_y=bucket_y,
                       address=bucket_address,
                       note=bucket_note,
                       lng=bucket_lng,
                       lat=bucket_lat,
                       point='POINT(%f %f)' % (bucket_x, bucket_y),
                       ).save()

        elif sheet_name_match:
            ws = wb[sheet_name]
            print(sheet_name)
            for row in range(3, ws.max_row+1):
                survey_date = ws['A' + str(row)].value
                bucket_id = ws['B' + str(row)].value
                if isinstance(survey_date, datetime) == False or \
                        bucket_dict.get(bucket_id) == None:
                    break

                area = ws['C' + str(row)].value
                village = ws['D' + str(row)].value
                if '里' not in village:
                    village = village + '里'

                egg_num = ws['E' + str(row)].value if isinstance(ws['E' + str(row)].value, int) == True else NO_DATA
                if egg_num == 0:
                    egypt_egg_num = 0
                    white_egg_num = 0
                else:
                    egypt_egg_num = ws['F' + str(row)].value if isinstance(ws['F' + str(row)].value, int) == True else NO_DATA
                    white_egg_num = ws['G'+ str(row)].value if isinstance(ws['G' + str(row)].value, int) == True else NO_DATA

                larvae_num = ws['H' + str(row)].value if isinstance(ws['H' + str(row)].value, int) == True else NO_DATA
                if larvae_num == 0:
                    egypt_larvae_num = 0
                    white_larvae_num = 0
                else:
                    egypt_larvae_num = ws['I' + str(row)].value if isinstance(ws['I' + str(row)].value, int) == True else NO_DATA
                    white_larvae_num = ws['J' + str(row)].value if isinstance(ws['J' + str(row)].value, int) == True else NO_DATA
                survey_note = ws['K' + str(row)].value if isinstance(ws['K' + str(row)].value, str) == True else '無'

                if([city, area, village] not in village_list):
                    village_list.append([city, area, village])

                BucketRecord(
                    bucket_id=bucket_id,
                    investigate_date=survey_date,
                    county=city,
                    town=area,
                    village=village,
                    # 卵數、埃及孵化卵數、白線孵化卵數
                    egg_count=egg_num,
                    egypt_egg_count=egypt_egg_num,
                    white_egg_count=white_egg_num,
                    # 孑孓、埃及幼蟲、白線幼蟲
                    larvae_count=larvae_num,
                    egypt_larvae_count=egypt_larvae_num,
                    white_larvae_count=white_larvae_num,
                    note=survey_note,
                ).save()
    wb.close()

for item in village_list:
    county = item[0]
    town = item[1]
    village = item[2]

    village_BucketRecords = BucketRecord.objects.filter(
            county=county
        ).filter(
            town=town
        ).filter(village=village)

    investigate_dates = []
    for record in village_BucketRecords:
        if(record.investigate_date not in investigate_dates):
            investigate_dates.append(record.investigate_date)

    for date in investigate_dates:
        date_of_village_BucketRecords = village_BucketRecords.filter(investigate_date=date)
        total_egg_count = sum(r.egg_count for r in date_of_village_BucketRecords)
        bucketes_has_egg = sum(r.egg_count > 0 for r in date_of_village_BucketRecords)
        positive_rate = bucketes_has_egg / len(date_of_village_BucketRecords)
        avg_egg_count = 10 * total_egg_count / len(date_of_village_BucketRecords)

        BucketStatistics(
            investigate_date=date,
            county=county,
            town=town,
            village=village,
            total_egg_count=total_egg_count,
            positive_rate=positive_rate,
            avg_egg_count=avg_egg_count
        ).save()


# FOR HEAT MAP
# for week_range_str in survey_dict.keys():
#     week_start = datetime.strptime(week_range_str.split("~")[0], "%Y-%m-%d")
#     week_str_list = list()
#     week_str_list.append(generate_week_str(week_start, 1))
#     week_str_list.append(generate_week_str(week_start, 2))
#     week_str_list.append(generate_week_str(week_start, 3))
#     week_str_list.append(generate_week_str(week_start, 4))

#     for city in survey_dict[week_range_str].keys():
#         for area in survey_dict[week_range_str][city].keys():
#             for village in survey_dict[week_range_str][city][area].keys():
#                 for bucket_id in survey_dict[week_range_str][city][area][village].keys():
#                     total_egg_num = 0
#                     for week_str in week_str_list:
#                         try:
#                             total_egg_num = total_egg_num + \
#                                 survey_dict[week_str][city][area][village][bucket_id]
#                         except:
#                             continue
#                     survey_dict[week_range_str][city][area][village][bucket_id]['avg_egg_num'] = int(total_egg_num / len(week_str_list))

# for week_range_str in survey_dict.keys():
    # with open('%s.json' % (week_range_str), 'w') as myfile:
        # json.dump(
            # survey_dict[week_range_str],
            # myfile,
            # indent=4,
            # ensure_ascii=False
        # )
